"""
Nium Playbook — Scraper & Data Layer (no Streamlit dependency)

Provides:
  scrape_dataset()        — Selenium + BeautifulSoup scrape for one dataset type
  transform_raw_to_wide() — Pivot long-format rows into wide DataFrame
  save_scraped_data()     — Write to data/ and scraped_data/ archive
  log_scrape_failures()   — Append failed-country details to logs/scrape_failures.log
  get_last_updated()      — Timestamp of most recent data file
"""

import json
import os
import re
import time
import unicodedata
from io import BytesIO
from pathlib import Path
from datetime import datetime

import pandas as pd

from config import SCRAPE_PAGE_TIMEOUT_SECONDS, SCRAPE_MAX_RETRIES, SCRAPE_RETRY_DELAY_SECONDS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR / "data"
ARCHIVE_DIR = BASE_DIR / "scraped_data"
DATA_DIR.mkdir(exist_ok=True)
ARCHIVE_DIR.mkdir(exist_ok=True)

FI_PATH     = DATA_DIR / "FI_data.xlsx"
NON_FI_PATH = DATA_DIR / "Non_FI_data.xlsx"

# ─── Country list (loaded from countries.txt — edit that file to add/remove) ─
_countries_file = BASE_DIR / "countries.txt"
COUNTRIES = [
    line.strip()
    for line in _countries_file.read_text(encoding="utf-8").splitlines()
    if line.strip()
]

# ─── URL slug overrides (countries whose Nium slug ≠ auto-derived slug) ──────
# Edit country_url_overrides.json to fix a country that shows up in "failed".
_overrides_file = BASE_DIR / "country_url_overrides.json"
URL_OVERRIDES: dict = json.loads(_overrides_file.read_text(encoding="utf-8")) if _overrides_file.exists() else {}

# ─── Internal helpers ─────────────────────────────────────────────────────────

def _format_country_url(name: str) -> str:
    """Convert a country display name to its Nium Playbook URL slug.

    Checks URL_OVERRIDES first for any country whose slug doesn't match the
    auto-derived form, then falls back to: strip accents → lowercase →
    remove punctuation → collapse spaces/hyphens.
    """
    if name in URL_OVERRIDES:
        return URL_OVERRIDES[name]
    # Decompose accented chars: ç→c, é→e, ü→u, ñ→n, etc.
    nfkd = unicodedata.normalize("NFKD", name.strip())
    ascii_name = nfkd.encode("ascii", "ignore").decode("ascii")
    slug = ascii_name.lower()
    # Strip everything except letters, digits, spaces, hyphens
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    # Collapse runs of whitespace/hyphens into a single hyphen
    slug = re.sub(r"[\s-]+", "-", slug)
    return slug.strip("-")

def _normalize_title(t):
    return (t.replace("Bank Account (ACH) (BANK)", "Bank Account (ACH)")
             .replace("Wallet (WALLET)", "Wallet")
             .strip())

def _normalize_tat(t):
    return t.split("/")[0].strip()

def _extract_pills(container):
    pills = container.find_all("span", class_=lambda c: c and "rounded" in c)
    if pills:
        return ", ".join(p.get_text(strip=True) for p in pills)
    return container.get_text(strip=True)


# ─── Scraper ──────────────────────────────────────────────────────────────────

def scrape_dataset(dataset_type, progress_bar, status_text):
    """
    Scrape playbook.nium.com for all countries.
    dataset_type: 'FI' or 'Non-FI'
    Returns: (raw_rows: list[dict], failed: list[str])
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup

    url_suffix = "/institutions" if dataset_type == "FI" else ""

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    raw_rows = []
    total    = len(COUNTRIES)
    failed   = []

    for idx, country_name in enumerate(COUNTRIES):
        pct = (idx + 1) / total
        progress_bar.progress(pct, text=f"Scraping {dataset_type}: {country_name} ({idx+1}/{total})")
        status_text.caption(f"⏳ {country_name}...")

        url = f"https://playbook.nium.com/country/{_format_country_url(country_name)}{url_suffix}"

        page_loaded = False
        for attempt in range(SCRAPE_MAX_RETRIES):
            try:
                driver.get(url)
                WebDriverWait(driver, SCRAPE_PAGE_TIMEOUT_SECONDS).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "payouts-details-table"))
                )
                page_loaded = True
                break
            except Exception:
                if attempt < SCRAPE_MAX_RETRIES - 1:
                    time.sleep(SCRAPE_RETRY_DELAY_SECONDS)
        if not page_loaded:
            failed.append(country_name)
            continue

        soup  = BeautifulSoup(driver.page_source, "html.parser")
        spans = soup.find_all("span", class_="payouts-details-table")

        for span in spans:
            try:
                raw_title = span.find(string=True, recursive=False)
                if not raw_title:
                    continue
                title = _normalize_title(raw_title.strip())

                currency, tat = "", ""
                small_tag = span.find("small")
                if small_tag:
                    lines = [l.strip() for l in small_tag.stripped_strings if l.strip()]
                    if lines:
                        currency = lines[0]
                    if len(lines) > 1:
                        tat = _normalize_tat(lines[1])

                modes_div = span.find("div", class_="hidden")
                modes     = modes_div.get_text(strip=True) if modes_div else ""

                parent_h3 = span.find_parent("h3")
                if not parent_h3:
                    continue
                body_div = parent_h3.find_next("div", id=lambda x: x and "accordion-collapse-body" in x)
                if not body_div:
                    continue

                # Key-value blocks
                payout_blocks = body_div.find_all(
                    "div", class_="overflow-hidden bg-white border mb-5 font-normal payouts-details-table"
                )
                for block in payout_blocks:
                    rows = block.find_all("div", class_="bg-gray-50 px-4 py-5 sm:grid sm:grid-cols-3 sm:gap-4 sm:px-6")
                    for row in rows:
                        key_tag   = row.find("dt")
                        value_tag = row.find("dd")
                        if not key_tag or not value_tag:
                            continue
                        key = key_tag.get_text(strip=True)
                        if key in ("Mandatory data requirements", "Supporting Documents"):
                            value = _extract_pills(value_tag)
                        else:
                            value = value_tag.get_text(strip=True)
                        raw_rows.append({
                            "Country": country_name, "Payment Mode": title,
                            "Currency": currency, "TAT": tat,
                            "Supported Modes": modes, "Key": key, "Value": value,
                        })

                # Tables (transaction limits)
                for table in body_div.find_all("table"):
                    table_title = "Transaction limit per end-user"
                    parent_div  = table.find_parent("div")
                    if parent_div:
                        dt_tag = parent_div.find_previous_sibling("dt") or parent_div.find("dt")
                        if not dt_tag:
                            outer_div = parent_div.find_parent("div")
                            if outer_div:
                                dt_tag = outer_div.find("dt")
                        if dt_tag:
                            table_title = dt_tag.get_text(strip=True)

                    headers = []
                    thead   = table.find("thead")
                    if thead:
                        headers = [th.get_text(strip=True) for th in thead.find_all("th") if th.get_text(strip=True)]

                    tbody = table.find("tbody")
                    if tbody:
                        for tr in tbody.find_all("tr"):
                            cells = tr.find_all(["th", "td"])
                            if not cells:
                                continue
                            row_label = cells[0].get_text(strip=True)
                            for ci, cell in enumerate(cells[1:]):
                                if ci < len(headers):
                                    raw_rows.append({
                                        "Country": country_name, "Payment Mode": title,
                                        "Currency": currency, "TAT": tat,
                                        "Supported Modes": modes,
                                        "Key":   f"{table_title} - {headers[ci]} - {row_label}",
                                        "Value": cell.get_text(strip=True),
                                    })
            except Exception:
                continue

    driver.quit()
    return raw_rows, failed


def transform_raw_to_wide(raw_rows):
    """Convert long-format key-value rows into wide-format DataFrame."""
    if not raw_rows:
        return pd.DataFrame()

    df_long = pd.DataFrame(raw_rows)
    id_cols = ["Country", "Payment Mode", "Currency", "TAT", "Supported Modes"]

    # Rename scraped Keys that collide with id_cols — pivot + reset_index
    # raises ValueError: "cannot insert X, already exists" otherwise
    conflicts = set(id_cols) & set(df_long["Key"].unique())
    if conflicts:
        df_long["Key"] = df_long["Key"].replace({k: f"{k} (Detail)" for k in conflicts})

    df_wide = df_long.pivot_table(
        index=id_cols, columns="Key", values="Value", aggfunc="first"
    ).reset_index()
    df_wide.columns.name = None

    other_cols = [c for c in df_wide.columns if c not in id_cols]
    preferred_order = [
        "Supported Modes 1", "Supported Currencies", "Network Participant",
        "Channels", "Cutoff & delivery timing", "Mandatory data requirements",
        "Supporting Documents", "Beneficiary Statement Narrative", "Proof of Payment", "Notes",
    ]
    ordered = [c for c in preferred_order if c in other_cols]
    remaining = sorted(c for c in other_cols if c not in ordered)
    return df_wide[id_cols + ordered + remaining]


def save_scraped_data(df, dataset_type):
    """Save to data/ (live) and scraped_data/ (date-stamped archive)."""
    today    = datetime.now().strftime("%Y-%m-%d")
    app_path = FI_PATH if dataset_type == "FI" else NON_FI_PATH
    df.to_excel(str(app_path), index=False)
    archive_path = ARCHIVE_DIR / f"{dataset_type}_{today}.xlsx"
    df.to_excel(str(archive_path), index=False)
    return str(app_path), str(archive_path)


def log_scrape_failures(failures_by_dataset: dict) -> str | None:
    """Append failed-country details to logs/scrape_failures.log.

    failures_by_dataset: {"FI": [...], "Non-FI": [...]}
    Returns the log file path, or None if there were no failures.
    """
    all_failed = {ds: countries for ds, countries in failures_by_dataset.items() if countries}
    if not all_failed:
        return None
    log_dir  = BASE_DIR / "logs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / "scrape_failures.log"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [f"\n{'='*60}", f"Scrape run: {timestamp}"]
    for ds, countries in all_failed.items():
        lines.append(f"\n[{ds}] {len(countries)} countries skipped:")
        for c in sorted(countries):
            lines.append(f"  - {c}")
    lines.append("")
    with open(log_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines))
    return str(log_path)


def create_formatted_excel(data, selected_cols, dataset_type):
    """Build a styled Excel workbook from a filtered DataFrame and return a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Nium Capabilities'

    blue_header   = '4472C4'
    thin_border   = Border(
        left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),  bottom=Side(style='thin', color='B4B4B4'),
    )
    header_border = Border(
        left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),  bottom=Side(style='medium', color='2F5496'),
    )

    ws.row_dimensions[1].height = 8
    ws.merge_cells('A2:E2')
    title_cell           = ws['A2']
    title_cell.value     = "Nium Payout Capability Matrix"
    title_cell.font      = Font(name='Segoe UI Semibold', size=14, bold=True, color='1A1A1A')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 6

    headers = ['#'] + selected_cols
    for col_idx, header in enumerate(headers, 1):
        cell           = ws.cell(row=4, column=col_idx)
        cell.value     = header
        cell.font      = Font(name='Segoe UI Semibold', size=11, bold=True, color='FFFFFF')
        cell.fill      = PatternFill(start_color=blue_header, end_color=blue_header, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border    = header_border
    ws.row_dimensions[4].height = 24
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'

    export_df = data[selected_cols].reset_index(drop=True)
    for row_idx, (_, row) in enumerate(export_df.iterrows()):
        excel_row         = row_idx + 5
        sn_cell           = ws.cell(row=excel_row, column=1)
        sn_cell.value     = row_idx + 1
        sn_cell.font      = Font(name='Segoe UI Semilight', size=11, color='333333')
        sn_cell.alignment = Alignment(horizontal='right', vertical='top')
        sn_cell.border    = thin_border
        for col_idx, col_name in enumerate(selected_cols):
            cell           = ws.cell(row=excel_row, column=col_idx + 2)
            val            = row[col_name]
            cell.value     = None if (pd.isna(val) or str(val) in ('nan', 'None', '')) else str(val)
            cell.font      = Font(name='Segoe UI Semilight', size=11, color='333333')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border    = thin_border

    ws.column_dimensions['A'].width = 5
    for col_idx, col_name in enumerate(selected_cols):
        col_letter = get_column_letter(col_idx + 2)
        max_len    = len(col_name)
        for ri in range(5, ws.max_row + 1):
            cv = ws.cell(row=ri, column=col_idx + 2).value
            if cv:
                max_len = max(max_len, min(len(str(cv)), 45))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    ws.freeze_panes = 'B5'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def get_last_updated():
    """Return human-readable timestamp of the FI data file, or None."""
    if FI_PATH.exists():
        ts = os.path.getmtime(str(FI_PATH))
        return datetime.fromtimestamp(ts).strftime("%d %b %Y, %I:%M %p")
    return None
